#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Excel转SQL导入工具
=================
将Excel格式的题库数据转换为Supabase SQL导入文件

使用方法：
  python excel_to_sql.py 题库.xlsx
  python excel_to_sql.py 题库.xlsx --output custom.sql

需要安装：pip install openpyxl
"""

import json
import argparse
from datetime import datetime
from pathlib import Path

try:
    from openpyxl import load_workbook
except ImportError:
    print("❌ 需要安装 openpyxl")
    print("   运行：pip install openpyxl")
    exit(1)


def excel_to_sql(excel_file: str, output_file: str = None, year: int = 2024):
    """将Excel转换为SQL"""
    
    print(f"🚀 Excel转SQL工具")
    print(f"{'='*70}\n")
    
    # 读取Excel
    print(f"📖 读取文件：{excel_file}")
    wb = load_workbook(excel_file)
    ws = wb.active
    
    # 读取表头
    headers = []
    for cell in ws[1]:
        headers.append(cell.value)
    
    print(f"✅ 表头：{headers}\n")
    
    # 读取数据
    questions = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row[0]:  # 跳过空行
            continue
        
        q = {
            'number': int(row[0]) if row[0] else 0,
            'type': row[1] or 'single',
            'content': row[2] or '',
            'options': {
                'A': row[3] or '',
                'B': row[4] or '',
                'C': row[5] or '',
                'D': row[6] or '',
                'E': row[7] or ''
            },
            'answer': row[8] or '',
            'explanation': row[9] or ''
        }
        
        questions.append(q)
    
    print(f"✅ 成功读取 {len(questions)} 道题\n")
    
    # 验证数据
    print("🔍 验证数据...")
    errors = []
    for i, q in enumerate(questions, 1):
        if not q['content']:
            errors.append(f"第{q['number']}题：题目内容为空")
        if not q['answer']:
            errors.append(f"第{q['number']}题：缺少答案")
        
        # 检查选项
        empty_options = [k for k, v in q['options'].items() if not v]
        if len(empty_options) > 2:  # 允许有2个空选项（有些题只有ABC）
            errors.append(f"第{q['number']}题：选项过少")
    
    if errors:
        print(f"⚠️  发现 {len(errors)} 个问题：")
        for error in errors[:10]:
            print(f"   - {error}")
        if len(errors) > 10:
            print(f"   ... 还有 {len(errors) - 10} 个问题")
        
        response = input("\n是否继续？(y/n): ")
        if response.lower() != 'y':
            print("❌ 已取消")
            return False
    else:
        print("✅ 数据验证通过\n")
    
    # 生成SQL
    print("📝 生成SQL文件...")
    
    if not output_file:
        output_file = f'import-{year}-questions-complete.sql'
    
    sql_parts = []
    
    # 文件头
    header = f"""-- ================================================================
-- 医考题库导入SQL - 从Excel生成
-- ================================================================
-- 年份：{year}
-- 题目总数：{len(questions)} 道
-- 生成时间：{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}
-- 数据来源：{excel_file}
-- ================================================================

-- 清理现有数据
DELETE FROM questions 
WHERE exam_type = '执业药师' 
  AND subject = '中药学综合知识与技能' 
  AND source_year = {year};

-- 批量插入
"""
    sql_parts.append(header)
    
    # 题型映射
    type_map = {
        'single': '最佳选择题',
        'match': '配伍选择题',
        'comprehensive': '综合分析题',
        'multiple': '多项选择题'
    }
    
    # 生成INSERT语句
    for q in questions:
        num = q['number']
        qtype = q['type']
        type_name = type_map.get(qtype, qtype)
        
        content = str(q['content']).replace("'", "''").replace("\\", "\\\\")
        explanation = str(q['explanation']).replace("'", "''").replace("\\", "\\\\")
        
        # 处理选项 - 只包含非空选项
        options_list = []
        for key in ['A', 'B', 'C', 'D', 'E']:
            value = q['options'].get(key, '')
            if value:
                options_list.append({'key': key, 'value': str(value)})
        
        options_json = json.dumps(options_list, ensure_ascii=False).replace("'", "''")
        
        sql = f"""
-- 第{num}题 - {type_name}
INSERT INTO questions (
  exam_type, subject, chapter, question_type, content, options,
  correct_answer, explanation, difficulty, knowledge_points,
  source_type, source_year, is_published
) VALUES (
  '执业药师',
  '中药学综合知识与技能',
  '综合知识',
  '{qtype}',
  '{content}',
  '{options_json}'::json,
  '{q["answer"]}',
  '{explanation}',
  2,
  ARRAY['综合知识'],
  '历年真题',
  {year},
  true
);
"""
        sql_parts.append(sql)
    
    # 验证查询
    footer = f"""
-- ================================================================
-- 验证导入结果
-- ================================================================
SELECT 
  question_type as "题型",
  COUNT(*) as "数量"
FROM questions 
WHERE source_year = {year}
GROUP BY question_type
ORDER BY 
  CASE question_type
    WHEN 'single' THEN 1
    WHEN 'match' THEN 2
    WHEN 'comprehensive' THEN 3
    WHEN 'multiple' THEN 4
  END;

SELECT COUNT(*) as "总数" FROM questions WHERE source_year = {year};
"""
    sql_parts.append(footer)
    
    # 保存文件
    sql_content = '\n'.join(sql_parts)
    with open(output_file, 'w', encoding='utf-8') as f:
        f.write(sql_content)
    
    print(f"✅ SQL文件生成成功：{output_file}")
    print(f"   文件大小：{len(sql_content) / 1024:.1f} KB")
    print(f"   题目数量：{len(questions)} 道\n")
    
    # 统计信息
    type_count = {}
    for q in questions:
        qtype = q['type']
        type_count[qtype] = type_count.get(qtype, 0) + 1
    
    print("📊 题型分布：")
    for qtype, count in sorted(type_count.items()):
        type_name = type_map.get(qtype, qtype)
        print(f"   - {type_name}: {count} 道")
    
    print(f"\n{'='*70}")
    print("💡 下一步：")
    print(f"   1. 打开 Supabase SQL 编辑器")
    print(f"   2. 复制粘贴 {output_file} 的内容")
    print(f"   3. 点击运行")
    print(f"{'='*70}\n")
    
    return True


def main():
    parser = argparse.ArgumentParser(description='Excel转SQL工具')
    parser.add_argument('input', help='输入Excel文件')
    parser.add_argument('--output', '-o', help='输出SQL文件')
    parser.add_argument('--year', '-y', type=int, default=2024, help='年份')
    
    args = parser.parse_args()
    
    if not Path(args.input).exists():
        print(f"❌ 文件不存在：{args.input}")
        return
    
    excel_to_sql(args.input, args.output, args.year)


if __name__ == '__main__':
    main()
